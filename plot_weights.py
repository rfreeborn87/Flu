#Let's do some flu stuff!
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import re
from nltk.tokenize import regexp_tokenize
import copy
import openpyxl
import os

"""This script imports an excel spreadsheet with animal weights over time.  It calculates the average change in weight per day for each group of animals, specified by the user. 
It plots both the weight change for individual animals as well as the average weight change for each group.  Plots are added to a new excel spreadsheet."""


file = pd.read_excel(input('Please enter the file which contains the original weight data.  Be sure to include path and extension (i.e. C:/Desktop/Rockwell_Lab/file.xlsx):\n'))
file.index = file['ID']
del file['ID']
#If a day's data hasn't been filled in yet, delete the column for that day.
new_df = copy.deepcopy(file).dropna(how = 'all', axis = 1).dropna(how = 'any')
#new_df.dropna(axis = 1)
if input('Do you have known uninfected animals?  (y/n) \n') == 'y':
    outliers = input('Enter the outliers\' IDs separated by a comma. \n').split(', ')
    for outlier in outliers:
        #print(outlier)
        new_df = new_df[new_df.index != outlier] 
    


#Build a list of days to serve as the x-axis for graph.  This should start at 0 and end at the last date with data entered.
count = 0
days = [count]
for column in range(1,len(new_df.columns)):
    count += 1
    days.append(count)
#Change the column names to be the day number instead of the date (the column names are dates in the original excel file).
new_df.columns = days[0:len(new_df.columns)]
#Make a new dataframe to calculate the weight percentage change.
percentage_df = copy.deepcopy(new_df)

#print(percentage_df.head())

#For each day, calculate the percent weight change
for day in percentage_df:
    if day == 0:
        continue
    for animal in percentage_df.index:
        percentage_df.loc[animal, day] = (percentage_df.loc[animal, day] - percentage_df.loc[animal, 0])/(percentage_df.loc[animal, 0]) * 100
for animal in percentage_df.index:
    percentage_df.loc[animal,0] = 0


#Figure out the treatment groups.  I should probably clean the data set to make it tidy instead of doing this, but that can wait.
identifiers = input('What are your groups?  Enter them as they appear in your excel file, separated by commas: ').split(', ')

#Make a list which will hold the averages for each group per day (i.e. [[day0 CS ave, day1 CS ave], [day0 TS ave, day1 TS ave]])
all_days = []
std_err = []
neg_err = []
#For each group, make an empty list which will hold the averages of a single day.
for group in identifiers:
    group_list = []
    group_errors = []
    group_neg_errors = []
    #For each day, create an empty list which will hold the values for each group.
    for day in percentage_df:
        day_vals = []
        #If the index of a value matches the group identifier, add that value to the list of values for that day.  This will make a new list for each group inside the outer list, day_vals.
        for index in percentage_df.index:
            if index[0:len(group)] == group:
                day_vals.append(percentage_df.loc[index, day])
        #Calculate the mean for each group on a specific day.
        day_average = np.mean(day_vals)
        #Calculate the standard error of the mean for each group on a specific day. 
        day_sem = (np.std(day_vals))
        day_neg_bar = 0
        #Append the mean to the list for a specific group.
        group_list.append(day_average)
        group_errors.append(day_sem)
        group_neg_errors.append(day_neg_bar)
    #Append the lists of averages to the master list, all_days.
    all_days.append(group_list)
    std_err.append(group_errors)
    neg_err.append(group_neg_errors)


#Make a new excel file in which we store the original data, data missing NaN values, and calculated weight change percentages.
new_file = input('What would you like your new file with graphs to be called?  Please enter the path, filename, and extension (i.e. C:/Desktop/Folder/filename.xlsx):\n')
writer = pd.ExcelWriter(new_file)
file.to_excel(writer, 'Original')
new_df.to_excel(writer, 'Cleaned')
percentage_df.to_excel(writer, 'Percentages')
writer.save()



#Transpose the dataframe so graphing is more simple.
percentage_df_for_plotting = percentage_df.transpose()
#Plot the weight change data for each animal.
plt.figure()
percentage_df_for_plotting.plot()#.legend(bbox_to_anchor = (1.04, 1), ncol = 3)
plt.xlabel('Day post-instillation')
plt.ylabel('Change in Mass (%)')
plt.legend(bbox_to_anchor = (1.05, 0.75), borderaxespad=0., ncol = 3)
#plt.savefig("individual_masses.png")
plt.savefig('individual_masses.png', bbox_inches='tight')


#Plot the weight change data for each group.
plt.figure()
fix, ax = plt.subplots()
for item in range(0,len(all_days)):
    #ax.plot(days, all_days[item], marker = 'o', label = identifiers[item])
    ax.errorbar(days, all_days[item], yerr = [neg_err[item],std_err[item]], marker = 'o', label = identifiers[item], capsize = 4)
plt.xlabel('Day post-instillation')
plt.ylabel('Change in Mass (%)')
ax.legend(loc = 3)
plt.savefig("averages.png", dpi = 100)

#Load the organized workbook into 'wb' which will be used to add graphs to new tabs.
wb = openpyxl.load_workbook(new_file)

#Create a new sheet in the Excel workbook.
ws = wb.create_sheet('individual_masses')
#Make an image object to be planted into Excel.
img = openpyxl.drawing.image.Image("individual_masses.png")
#Add the image.
ws.add_image(img, 'A1')


#Create a new sheet in the Excel workbook.
ws2 = wb.create_sheet('averages')
#Make an image object to be planted into Excel.
img2 = openpyxl.drawing.image.Image('averages.png')
#Add the image.
ws2.add_image(img2, 'A1')

#Resave the organized workbook with the added images.	
wb.save(new_file)
wb.close()

os.startfile(new_file)